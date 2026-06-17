import logging
import os
from typing import Dict, Tuple

import pandas as pd
from openpyxl import load_workbook

from model import LoadProfile, NodepoolConfig, AKSInfrastructure, APIConfig
from config import (
    DEFAULT_USERS, DEFAULT_INTERACTIONS_PER_USER_DAY,
    DEFAULT_INPUT_TOKENS, DEFAULT_OUTPUT_TOKENS,
    DEFAULT_WORKING_DAYS, DEFAULT_OFFICE_HOURS,
    DEFAULT_PEAK_HOURS, DEFAULT_CONCURRENT_RATIO,
    DEFAULT_PEAK_MULTIPLIER,
    DEFAULT_SYSTEM_VM, DEFAULT_SYSTEM_PRICE, DEFAULT_SYSTEM_NODES,
    DEFAULT_IDEAL_GPU_VM, DEFAULT_IDEAL_GPU_PRICE, DEFAULT_IDEAL_THROUGHPUT,
    DEFAULT_ECO_GPU_VM, DEFAULT_ECO_GPU_PRICE, DEFAULT_ECO_THROUGHPUT,
    DEFAULT_STORAGE_IDEAL, DEFAULT_STORAGE_ECO,
    DEFAULT_LB, DEFAULT_MONITOR_IDEAL, DEFAULT_MONITOR_ECO,
    DEFAULT_ACR_IDEAL, DEFAULT_ACR_ECO,
    DEFAULT_API_MODEL, DEFAULT_API_INPUT_PRICE, DEFAULT_API_OUTPUT_PRICE,
    DEFAULT_EUR_USD,
    DEFAULT_GPU_UTILIZATION, DEFAULT_SAFETY_FACTOR,
)

logger = logging.getLogger(__name__)


def _int(m: dict, key: str, default: int) -> int:
    v = m.get(key, default)
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return default


def _float(m: dict, key: str, default: float) -> float:
    v = m.get(key, default)
    try:
        return float(str(v))
    except (ValueError, TypeError):
        return default


def _str(m: dict, key: str, default: str) -> str:
    v = m.get(key, default)
    return str(v).strip() if v else default


def build_default_data() -> dict:
    load_profile = LoadProfile(
        users=DEFAULT_USERS,
        interactions_per_user_day=DEFAULT_INTERACTIONS_PER_USER_DAY,
        input_tokens_per_interaction=DEFAULT_INPUT_TOKENS,
        output_tokens_per_interaction=DEFAULT_OUTPUT_TOKENS,
        working_days_per_month=DEFAULT_WORKING_DAYS,
        office_hours_per_day=DEFAULT_OFFICE_HOURS,
        peak_hours_per_day=DEFAULT_PEAK_HOURS,
        concurrent_user_ratio=DEFAULT_CONCURRENT_RATIO,
        peak_multiplier=DEFAULT_PEAK_MULTIPLIER,
    )

    infra_ideal = AKSInfrastructure(name="LLM on AKS (Ideal UX)")
    infra_ideal.system_nodepool = NodepoolConfig(
        vm_type=DEFAULT_SYSTEM_VM, base_office_nodes=DEFAULT_SYSTEM_NODES,
        price_per_hour=DEFAULT_SYSTEM_PRICE,
    )
    infra_ideal.inference_nodepool = NodepoolConfig(
        vm_type=DEFAULT_IDEAL_GPU_VM,
        base_office_nodes=3, peak_nodes=10, off_hours_nodes=1,
        price_per_hour=DEFAULT_IDEAL_GPU_PRICE,
    )
    infra_ideal.throughput_tok_s_per_pod = DEFAULT_IDEAL_THROUGHPUT
    infra_ideal.gpu_utilization = 0.75
    infra_ideal.safety_factor = 1.0
    infra_ideal.base_replicas = 3
    infra_ideal.peak_replicas = 10
    infra_ideal.off_hours_replicas = 1
    infra_ideal.storage_cost_per_month = DEFAULT_STORAGE_IDEAL
    infra_ideal.lb_cost_per_month = DEFAULT_LB
    infra_ideal.monitor_cost_per_month = DEFAULT_MONITOR_IDEAL
    infra_ideal.acr_cost_per_month = DEFAULT_ACR_IDEAL

    infra_economica = AKSInfrastructure(name="LLM on AKS (Economy UX)")
    infra_economica.system_nodepool = NodepoolConfig(
        vm_type=DEFAULT_SYSTEM_VM, base_office_nodes=DEFAULT_SYSTEM_NODES,
        price_per_hour=DEFAULT_SYSTEM_PRICE,
    )
    infra_economica.inference_nodepool = NodepoolConfig(
        vm_type=DEFAULT_ECO_GPU_VM,
        base_office_nodes=5, peak_nodes=20, off_hours_nodes=1,
        price_per_hour=DEFAULT_ECO_GPU_PRICE,
    )
    infra_economica.throughput_tok_s_per_pod = DEFAULT_ECO_THROUGHPUT
    infra_economica.gpu_utilization = 0.75
    infra_economica.safety_factor = 1.0
    infra_economica.base_replicas = 5
    infra_economica.peak_replicas = 20
    infra_economica.off_hours_replicas = 1
    infra_economica.storage_cost_per_month = DEFAULT_STORAGE_ECO
    infra_economica.lb_cost_per_month = DEFAULT_LB
    infra_economica.monitor_cost_per_month = DEFAULT_MONITOR_ECO
    infra_economica.acr_cost_per_month = DEFAULT_ACR_ECO

    api_config = APIConfig(
        name="API Azure OpenAI",
        model=DEFAULT_API_MODEL,
        input_price_per_1m_tokens_usd=DEFAULT_API_INPUT_PRICE,
        output_price_per_1m_tokens_usd=DEFAULT_API_OUTPUT_PRICE,
        eur_usd_rate=DEFAULT_EUR_USD,
    )

    return {
        "load_profile": load_profile,
        "infra_ideal": infra_ideal,
        "infra_economica": infra_economica,
        "api_config": api_config,
        "comparativa": pd.DataFrame(),
    }


def extract_all(path: str = "") -> dict:
    if not path or not os.path.exists(path):
        logger.info("Excel not found: %s. Using built-in defaults.", path)
        return build_default_data()

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active

        m = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    m[cell.coordinate] = cell.value

        wb.close()
    except Exception as e:
        logger.error("Error reading Excel %s: %s", path, e)
        return build_default_data()

    if "B9" not in m:
        logger.warning("Unknown Excel format in %s. Using defaults.", path)
        return build_default_data()

    load_profile = LoadProfile(
        users=_int(m, "B9", DEFAULT_USERS),
        interactions_per_user_day=_int(m, "B10", DEFAULT_INTERACTIONS_PER_USER_DAY),
        input_tokens_per_interaction=_int(m, "B11", DEFAULT_INPUT_TOKENS),
        output_tokens_per_interaction=_int(m, "B12", DEFAULT_OUTPUT_TOKENS),
        working_days_per_month=_int(m, "B13", DEFAULT_WORKING_DAYS),
        office_hours_per_day=_int(m, "B14", DEFAULT_OFFICE_HOURS),
        peak_hours_per_day=_float(m, "B16", DEFAULT_PEAK_HOURS),
        concurrent_user_ratio=0.20,
        peak_multiplier=1.0,
    )

    infra_ideal = AKSInfrastructure(name="LLM on AKS (Ideal UX)")
    infra_ideal.system_nodepool = NodepoolConfig(
        vm_type=_str(m, "G12", DEFAULT_SYSTEM_VM),
        base_office_nodes=_int(m, "G13", DEFAULT_SYSTEM_NODES),
        price_per_hour=_float(m, "B36", DEFAULT_SYSTEM_PRICE),
    )
    infra_ideal.inference_nodepool = NodepoolConfig(
        vm_type=_str(m, "G14", DEFAULT_IDEAL_GPU_VM),
        base_office_nodes=_int(m, "G15", 3),
        peak_nodes=_int(m, "G16", 43),
        off_hours_nodes=_int(m, "G17", 1),
        price_per_hour=_float(m, "B37", DEFAULT_IDEAL_GPU_PRICE),
    )
    infra_ideal.throughput_tok_s_per_pod = _float(m, "K9", DEFAULT_IDEAL_THROUGHPUT)
    infra_ideal.gpu_utilization = _float(m, "K11", DEFAULT_GPU_UTILIZATION)
    infra_ideal.safety_factor = _float(m, "K12", DEFAULT_SAFETY_FACTOR)
    infra_ideal.base_replicas = _int(m, "G15", 3)
    infra_ideal.peak_replicas = _int(m, "G16", 43)
    infra_ideal.off_hours_replicas = _int(m, "G17", 1)
    infra_ideal.storage_cost_per_month = _float(m, "B38", DEFAULT_STORAGE_IDEAL)
    infra_ideal.lb_cost_per_month = _float(m, "B39", DEFAULT_LB)
    infra_ideal.monitor_cost_per_month = _float(m, "B40", DEFAULT_MONITOR_IDEAL)
    infra_ideal.acr_cost_per_month = _float(m, "B41", DEFAULT_ACR_IDEAL)

    infra_economica = AKSInfrastructure(name="LLM on AKS (Economy UX)")
    infra_economica.system_nodepool = NodepoolConfig(
        vm_type=_str(m, "G25", DEFAULT_SYSTEM_VM),
        base_office_nodes=_int(m, "G26", DEFAULT_SYSTEM_NODES),
        price_per_hour=_float(m, "G36", DEFAULT_SYSTEM_PRICE),
    )
    infra_economica.inference_nodepool = NodepoolConfig(
        vm_type=DEFAULT_ECO_GPU_VM,
        base_office_nodes=_int(m, "G27", 3),
        peak_nodes=_int(m, "G28", 124),
        off_hours_nodes=_int(m, "G29", 1),
        price_per_hour=_float(m, "G37", DEFAULT_ECO_GPU_PRICE),
    )
    infra_economica.throughput_tok_s_per_pod = _float(m, "K10", DEFAULT_ECO_THROUGHPUT)
    infra_economica.gpu_utilization = _float(m, "K11", DEFAULT_GPU_UTILIZATION)
    infra_economica.safety_factor = _float(m, "K12", DEFAULT_SAFETY_FACTOR)
    infra_economica.base_replicas = _int(m, "G27", 5)
    infra_economica.peak_replicas = _int(m, "G28", 124)
    infra_economica.off_hours_replicas = _int(m, "G29", 1)
    infra_economica.storage_cost_per_month = _float(m, "G38", DEFAULT_STORAGE_ECO)
    infra_economica.lb_cost_per_month = _float(m, "G39", DEFAULT_LB)
    infra_economica.monitor_cost_per_month = _float(m, "G40", DEFAULT_MONITOR_ECO)
    infra_economica.acr_cost_per_month = _float(m, "G41", DEFAULT_ACR_ECO)

    api_config = APIConfig(
        name="API Azure OpenAI",
        model=_str(m, "K36", DEFAULT_API_MODEL),
        input_price_per_1m_tokens_usd=_float(m, "K37", DEFAULT_API_INPUT_PRICE),
        output_price_per_1m_tokens_usd=_float(m, "K38", DEFAULT_API_OUTPUT_PRICE),
        eur_usd_rate=_float(m, "K39", DEFAULT_EUR_USD),
    )

    logger.info(
        "Excel loaded: %d users, %d int/user/day, %d/%d tok in/out, "
        "ideal=%d/%d/%d nodes, eco=%d/%d/%d nodes, api=%.2f/%.2f $/1M tok",
        load_profile.users, load_profile.interactions_per_user_day,
        load_profile.input_tokens_per_interaction, load_profile.output_tokens_per_interaction,
        infra_ideal.inference_nodepool.base_office_nodes,
        infra_ideal.inference_nodepool.peak_nodes,
        infra_ideal.inference_nodepool.off_hours_nodes,
        infra_economica.inference_nodepool.base_office_nodes,
        infra_economica.inference_nodepool.peak_nodes,
        infra_economica.inference_nodepool.off_hours_nodes,
        api_config.input_price_per_1m_tokens_usd,
        api_config.output_price_per_1m_tokens_usd,
    )

    return {
        "load_profile": load_profile,
        "infra_ideal": infra_ideal,
        "infra_economica": infra_economica,
        "api_config": api_config,
        "comparativa": pd.DataFrame(),
    }
